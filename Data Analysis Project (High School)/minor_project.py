"""
Version 6 (Final Version):

- Introduced even more functionality to the program including, removing all outliers
and writing to a new sheet and filling in all blanks and writing to a new sheet, graphing
data for specific ranges or for the whole column and determing what values are outside
the accepted range for a substance.
- Also added an option to quit the program
- Program has also been split into seperate files to increase maintainability and reusability of code
and functions as well as strings.
"""

# Importing all functions
import sys # Used for finding size of variables
import os # Used for saving file
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis

# '*' Imports all functions from specified file
from project_functions import *
from project_strings import *

# Ensures program can run correctly on different devices from different users
folder_directory = os.path.dirname(__file__)
os.chdir(folder_directory)

# Importing workbook
workbook_name = 'project_data.xlsx'
workbook = load_workbook(workbook_name)

# Storing sheets in varaible names
summer_sheet = workbook['original_data_summer']
autumn_sheet = workbook['original_data_autumn']
both_sheet = workbook['original_data_both']
test_sheet = workbook['test_data']

run_main = True
while run_main:
    print("""
--------------------------------------------------------------------------------------------------------
Data Analysis of Sydney Particle Study Dataset:
--------------------------------------------------------------------------------------------------------""")

    valid_sheet_choice = False
    while not(valid_sheet_choice): # Getting sheet the user wants to analyse
        sheet_choice = input("""
Enter the sheet number you would like to analyse data from:
1) Summer
2) Autumn
3) Both Summer and Autumn
4) Test data
(Enter corresponding number or enter 'q' to quit the program)
--------------------------------------------------------------------------------------------------------
> """)
        print('--------------------------------------------------------------------------------------------------------')

        # Sheets for only summer data, only autumn data and both
        if sheet_choice == '1':
            original_sheet = summer_sheet
            original_sheet_name = 'original_data_summer'
            sheet_name = 'data_summer'
            valid_sheet_choice = True
        elif sheet_choice == '2':
            original_sheet = autumn_sheet
            original_sheet_name = 'original_data_autumn'
            sheet_name = 'data_autumn'
            valid_sheet_choice = True
        elif sheet_choice == '3':
            original_sheet = both_sheet
            original_sheet_name = 'original_data_both'
            sheet_name = 'data_both'
            valid_sheet_choice = True
        elif sheet_choice == '4':
            original_sheet = test_sheet
            original_sheet_name = 'test_sheet'
            sheet_name = 'data_test'
            valid_sheet_choice = True
        elif sheet_choice.lower() == 'q':
            run_main = False
            break
        else:
            print("Invalid sheet choice")

    if run_main == False:
        continue

    headings = get_headings(original_sheet) # list of headings
    column_options = create_column_options_string(headings) # string with all column options used throughout program
    units = get_units(headings)
    # Units = ['', '', 'ppb', '%', 'bsp', 'ppb', 'ppb', 'ppb', 'ppb', 'µg/m≥', '∞', 'ppb', 'W/m2', '∞C', '∞', 'm/s']

    choice = input("""{}
--------------------------------------------------------------------------------------------------------
> """.format(options))
    print('--------------------------------------------------------------------------------------------------------')

    space_for_graph = -13 # Shifts new graphs down so all are aligned and have their own space
    keep_running = True

    while choice and keep_running: # Runs until user enters 13 (end command) or invalid input is given
        need_column_num = ('1', '2', '3', '4', '5', '6', '7')
        choices_eight_to_twelve = ('8', '9', '10', '11', '12')

        if choice in need_column_num: # Options 1 - 7
            valid_col = False
            while not(valid_col):
                try:
                    column_num = int(input("""{}
--------------------------------------------------------------------------------------------------------
> """.format(column_options))) + 2
                    print('--------------------------------------------------------------------------------------------------------')
                    total_columns = original_sheet.max_column - 2 # int
                    if column_num <= total_columns + 2 and column_num >= 3:
                        valid_col = True
                    else:
                        print("Invalid input")
                except ValueError:
                    print("Invalid input")
            
            
            substance = headings[column_num - 1]
            substance_units = units[column_num - 1]
            sorted_col = sort_column(column_num, original_sheet) # list of floats

            if len(sorted_col) == 0: # Checking for columns with no data
                print(f"No data available for {substance}")

            elif choice == '1': # Mean
                mean = get_mean(column_num, original_sheet, sorted_col)
                print(f"Mean of {substance} is {mean} {substance_units}")

            elif choice == '2': # Min, Max, Range
                minimum, maximum, range_data = get_min_max(column_num, original_sheet, sorted_col)
                print(f"Maximum of {substance} is {maximum} {substance_units}, minimum is {minimum} {substance_units} and range of data is {range_data} {substance_units}")
            
            elif choice == '3': # Modes
                modes = get_modes(column_num, original_sheet, sorted_col)
                if len(modes) > 1:
                    string = f"Modes of {substance} are "
                else:
                    string = f"Mode of {substance} is "

                for mode in modes:
                    string += f"{str(mode)} {substance_units}, "
                
                print(string[:-2])
            
            elif choice == '4': # q1, median, q3, IQR
                q1, median, q3, IQR = get_quartiles(column_num, original_sheet, sorted_col)
                print(f"Quartiles of {substance} are Q1 = {q1} {substance_units}, median (Q2) = {median} {substance_units} and Q3 = {q3} {substance_units}.\nIQR = {IQR} {substance_units}.")
            
            elif choice == '5': # Variance and standard deviation
                var, dev = get_variance_and_deviation(column_num, original_sheet, sorted_col)
                print(f"Variange of {substance} is {round(var, 3)} {substance_units} and standard deviation is {round(dev, 3)} {substance_units}.")

            elif choice == '6': # Summary stats
                summary_stats = get_summary_stats(column_num, original_sheet, substance, sorted_col, substance_units)
                print(summary_stats)
                print('--------------------------------------------------------------------------------------------------------')
                print("Would you like to write this information to a text file (y/n)")
                print('--------------------------------------------------------------------------------------------------------')
                write_to_file = input("> ")
                if write_to_file.lower() == 'y': # Writing summary stats to a file
                    print('--------------------------------------------------------------------------------------------------------')
                    print("Would you like to write to a (1) new file or (2) existing file? (Enter corresponding number)")
                    print('--------------------------------------------------------------------------------------------------------')
                    new_or_existing = int(input("> "))
                    print('--------------------------------------------------------------------------------------------------------')
                    print("Enter the name of the text file:")
                    print('--------------------------------------------------------------------------------------------------------')
                    txt_file_name = input("> ")
                    print('--------------------------------------------------------------------------------------------------------')
                    #final_file_name = '/Users/42587/Library/CloudStorage/OneDrive-StJosephsCollege/Year 12/SDD/Minor Project/' + txt_file_name
                    final_file_name = txt_file_name
                    # Might need to change if used on a differnet computer
                    with open(final_file_name, 'a') as txt_file:
                        txt_file.write(summary_stats)
                    
                    print(f"Successfully wrote the summary statistics to the file {txt_file_name}")

                elif write_to_file.lower() == 'n': # User doesn't want to write summary stats to a file
                    pass
                else:
                    print("Invalid input.")

            else: # Z-score for certain value
                print("Enter the value you would like to determine a z-score for")
                value = float(input("> "))
                z_score = calc_z_score(value, column_num, original_sheet, sorted_col)
                print('--------------------------------------------------------------------------------------------------------')
                if z_score > 0:
                    print(f"Z-score for {value} is {z_score}. Thus the data point {value} falls {z_score} standard deviations above the mean.")
                elif z_score < 0:
                    print(f"Z-score for {value} is {z_score}. Thus the data point {value} falls {abs(z_score)} standard deviations below the mean.")
                else:
                    print(f"Z-score for {value} is {z_score}. Thus the data point {value} is equal to the mean.")

            print('--------------------------------------------------------------------------------------------------------')

        elif choice in choices_eight_to_twelve: # Options 8 - 12
            if choice == '8':
                outlier_string = get_outlier_string(original_sheet, headings)
                print(outlier_string)
                print('--------------------------------------------------------------------------------------------------------')

            elif choice == '9':
                ### Removes outliers from original data, copies new data to new file, ###
                ### then finds avg or median of new data and fills in all blanks in the new ###
                ### data (including old gaps and new ones created when outliers were removed) ###
                ### then copies the final data without outliers and filled in to new sheet ###

                ### REMOVING OUTLIERS AND CREATING SHEET WITHOUT OUTLIERS ###
                no_outliers_sheet_name = create_no_outliers_sheet(sheet_name, workbook, original_sheet, original_sheet_name)
                no_outliers_sheet = workbook[no_outliers_sheet_name]
                
                ### FILLING IN BLANKS AND CREATING SHEET WITHOUT OUTLIERS AND WITH BLANKS FILLED IN ###
                worked, message = create_no_blanks_sheet(sheet_name, workbook, original_sheet, original_sheet_name, no_outliers_sheet)
                if worked:
                    print(message)
                else:
                    print("New sheet has not been created.")

            elif choice == '10' or choice == '11': # Graphing data
                print("""Creating scatterplot and line of best fit for specified column:
--------------------------------------------------------------------------------------------------------""")
                valid_col = False
                while not(valid_col):
                    try:
                        column_num = int(input("""{}
--------------------------------------------------------------------------------------------------------
> """.format(column_options))) + 2
                        total_columns = original_sheet.max_column - 2 # int
                        if column_num <= total_columns + 2 and column_num >= 3:
                            valid_col = True
                        else:
                            print("Invalid input")
                    except ValueError:
                        print("Invalid input")

                print('--------------------------------------------------------------------------------------------------------')
                total_rows = original_sheet.max_row

                if choice == '11': # Graphing certain time period
                    valid_range = False
                    while not(valid_range):
                        print("Enter range of rows you would like to graph in format: starting row, ending row")
                        start_str, end_str = input("> ").split(",")
                        print('--------------------------------------------------------------------------------------------------------')
                        start = int(start_str)
                        end = int(end_str)
                        if start < 2 or end > total_rows or start > end:
                            print("Invalid start and end points")
                        else:
                            valid_range = True
                else:
                    start = 2
                    end = total_rows
                
                column_name_old = headings[column_num - 1]
                column_name = column_name_old.replace("[", "").replace("]", "")
                starting_date = f"{original_sheet.cell(row=start, column=1).value}"

                # Creating line graph
                c1 = LineChart()
                c1.title = column_name
                c1.style = 13
                c1.y_axis.title = column_name
                c1.x_axis.title = f'Hours since {starting_date}'
                data = Reference(original_sheet, min_col=column_num, min_row=start, max_row=end) # Crating data for graph
                c1.add_data(data, titles_from_data=False)
                space_for_graph += 15
                new_cell = f"S{str(space_for_graph)}"
                original_sheet.add_chart(c1, anchor=new_cell) # Adding graph to sheet
                workbook.save('project_data.xlsx') # Saving file so that graph will appear in the sheet

                print("Successfully created graph!")
                print('--------------------------------------------------------------------------------------------------------')
            
            elif choice == '12': # determining values outside acceptable range
                valid_col = False
                while not(valid_col):
                    try:
                        column_num = int(input("""{}
--------------------------------------------------------------------------------------------------------
> """.format(column_options))) + 2
                        total_columns = original_sheet.max_column - 2 # int
                        if column_num <= total_columns + 2 and column_num >= 3:
                            valid_col = True
                        else:
                            print("Invalid input")
                    except ValueError:
                        print("Invalid input")

                print('--------------------------------------------------------------------------------------------------------')
                total_rows = original_sheet.max_row
                print("Input acceptable range for substance in the format: min, max")
                print('--------------------------------------------------------------------------------------------------------')
                minimum, maximum = input("> ").split(",")
                print('--------------------------------------------------------------------------------------------------------')
                sorted_col = sort_column(column_num, original_sheet)

                # Finds values outside acceptable range and returns as a string
                outside_range_string = determine_values_outside_range(sorted_col, maximum, minimum, headings, column_num, total_rows)
                
                print(outside_range_string)
                print('--------------------------------------------------------------------------------------------------------')
            
            else:
                print("Invalid input") 

        elif choice.lower() == 'q': # User wants to quit back to main menu
            print("Transferring back to main menu...")
            print('--------------------------------------------------------------------------------------------------------')
            keep_running = False
            continue 

        else:
            print("Invalid input.")
            print('--------------------------------------------------------------------------------------------------------')

        # Allows users to continually enter new actions from the options menu
        choice = input("""{}
--------------------------------------------------------------------------------------------------------
> """.format(options))
        print('--------------------------------------------------------------------------------------------------------')

print("\nExiting program...\n")