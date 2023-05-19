"""
Version 5:

- Introduced new functions and capabilities of program including getting variance,
standard deviation and finding outliers in the data

Note: Most filenames have been changed since this version was created
"""

from openpyxl import load_workbook
from openpyxl import Workbook

from openpyxl.chart.axis import DateAxis
from novice_functions import sort_column
from novice_functions import get_mean
from novice_functions import get_min_max
from novice_functions import get_modes
from novice_functions import get_middle_of_list
from novice_functions import get_quartiles
from novice_functions import check_for_outliers
from novice_functions import get_headings
from novice_functions import get_summary_stats
from novice_functions import get_variance_and_deviation
from novice_functions import create_column_options_string
from novice_strings import options

workbook = load_workbook("novice_data.xlsx")
summer_sheet = workbook['original_data_summer']
autumn_sheet = workbook['original_data_autumn']
both_sheet = workbook['original_data_both']
test_sheet = workbook['test_data']

print("""
----------------------------------------------------------------------------------------------
Data Analysis of Sydney Particle Study Dataset:
----------------------------------------------------------------------------------------------""")

sheet_choice = int(input("""
Would you like to analyse data collected in:
1) Summer
2) Autumn
3) Both Summer and Autumn
4) Test data
(Enter corresponding number)
----------------------------------------------------------------------------------------------
> """))
print("----------------------------------------------------------------------------------------------")

# Sheets for only summer data, only autumn data and both
if sheet_choice == 1:
    original_sheet = summer_sheet
    original_sheet_name = 'original_data_summer'
    sheet_name = 'data_summer'
elif sheet_choice == 2:
    original_sheet = autumn_sheet
    original_sheet_name = 'original_data_autumn'
    sheet_name = 'data_autumn'
elif sheet_choice == 3:
    original_sheet = both_sheet
    original_sheet_name = 'original_data_both'
    sheet_name = 'data_both'
elif sheet_choice == 4:
    original_sheet = test_sheet
    original_sheet_name = 'test_sheet'
    sheet_name = 'data_test'
else:
    print("Invalid sheet choice")

headings = get_headings(original_sheet) # list of headings
column_options = create_column_options_string(headings) # string with all column options used throughout program

choice = int(input("""{}
----------------------------------------------------------------------------------------------
> """.format(options)))
print("----------------------------------------------------------------------------------------------")

space_for_graph = -13 # Shifts new graphs down so all are aligned and have their own space

while choice: # Runs until user enters 13 (end command) or invalid input is given
    need_column_num = [1, 2, 3, 4, 5, 6, 7]
    if choice in need_column_num: # Options 1 - 5
        column_num = int(input("""{}
----------------------------------------------------------------------------------------------
> """.format(column_options))) + 2
        print("----------------------------------------------------------------------------------------------")
        substance = headings[column_num - 1]

        sorted_col = sort_column(column_num, original_sheet)
        if len(sorted_col) == 0: # Checking for columns with no data
            print(f"No data available for {substance}")

        elif choice == 1: # Mean
            mean = get_mean(column_num, original_sheet)
            print(f"Mean of {substance} is {mean}")

        elif choice == 2: # Min, Max, Range
            minimum, maximum, range_data = get_min_max(column_num, original_sheet)
            print(f"Maximum of {substance} is {maximum}, minimum is {minimum} and range of data is {range_data}")
        
        elif choice == 3: # Modes
            modes = get_modes(column_num, original_sheet)
            if len(modes) > 1:
                string = f"Modes of {substance} are "
            else:
                string = f"Mode of {substance} is "

            for mode in modes:
                string += (str(mode) + " ")
            print(string)
        
        elif choice == 4: # q1, median, q3, IQR
            q1, median, q3, IQR = get_quartiles(column_num, original_sheet)
            print(f"Quartiles of {substance} are Q1 = {q1}, median (Q2) = {median} and Q3 = {q3}.\nIQR = {IQR}")
        
        elif choice == 5: # Variance and standard deviation
            var, dev = get_variance_and_deviation(column_num, original_sheet)
            print(f"Variange of {substance} is {round(var, 3)} and standard deviation is {round(dev, 3)}.")

        elif choice == 6: # Summary stats
            summary_stats = get_summary_stats(column_num, original_sheet, substance)
            print("----------------------------------------------------------------------------------------------")
            print(summary_stats)
        else: # Z-score for certain value
            print("Enter the value you would like to determine a z-score for")
            value = float(input("> "))
            mean = get_mean(column_num, original_sheet)
            var, standard_deviation = get_variance_and_deviation(column_num, original_sheet)
            z_score = round((value - mean)/standard_deviation, 2)
            if z_score > 0:
                print(f"Z-score for {value} is {z_score}. Thus the data point {value} falls {z_score} standard deviations above the mean.")
            elif z_score < 0:
                print(f"Z-score for {value} is {z_score}. Thus the data point {value} falls {abs(z_score)} standard deviations below the mean.")
            else:
                print(f"Z-score for {value} is {z_score}. Thus the data point {value} is equal to the mean.")

        print("----------------------------------------------------------------------------------------------")

    else: # Option 8
        if choice == 8:
            total_columns = original_sheet.max_column # int
            outliers_string = "" # increased for each column processed in loop below
            cur_column = 3
            while cur_column <= total_columns:
                col_heading = headings[cur_column - 1] # string

                sorted_col = sort_column(cur_column, original_sheet)
                if len(sorted_col) == 0: # Columns that have no data
                    outliers_string += f"{col_heading} has no data\n\n"

                else: # Finding outliers
                    cur_col_outliers = check_for_outliers(cur_column, original_sheet) # list of floats

                    if len(cur_col_outliers) == 0: # Columns that have no outliers
                        outliers_string += f"{col_heading} has no outliers\n\n"
                    else: # Columns that have outliers - making a string adding all outliers
                        total_num_outliers = len(cur_col_outliers)
                        cur_col_string = f"{col_heading} has {total_num_outliers} outliers include"
                        for outlier in cur_col_outliers:
                            cur_col_string += f", {str(outlier)}"
                        
                        outliers_string += f"{cur_col_string}\n\n"

                cur_column += 1

            print(outliers_string)
            print("----------------------------------------------------------------------------------------------")
        